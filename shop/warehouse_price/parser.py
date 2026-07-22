import csv
from pathlib import Path
from typing import Iterator, List, Tuple

from shop.warehouse_price.columns import build_column_indexes
from shop.warehouse_price.normalize import cell_to_str, parse_price, parse_quantity
from shop.warehouse_price.types import ParsedPriceRow


def _read_rows_xlsx(path: Path) -> List[Tuple[int, List]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            'Для Excel (.xlsx) установите openpyxl: pip install openpyxl'
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows: List[Tuple[int, List]] = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        rows.append((row_idx, list(row)))
    workbook.close()
    return rows


def _read_rows_csv(path: Path) -> List[Tuple[int, List]]:
    raw = path.read_bytes()
    text = None
    encoding = 'utf-8-sig'
    for enc in ('utf-8-sig', 'utf-8', 'cp1251'):
        try:
            text = raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode('cp1251', errors='replace')

    first_line = text.splitlines()[0] if text.splitlines() else ''
    delimiter = ';' if first_line.count(';') >= first_line.count(',') else ','

    rows: List[Tuple[int, List]] = []
    for idx, row in enumerate(csv.reader(text.splitlines(), delimiter=delimiter), start=1):
        rows.append((idx, row))
    return rows


def read_sheet_rows(file_path: str) -> List[Tuple[int, List]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in ('.xlsx', '.xlsm'):
        return _read_rows_xlsx(path)
    if suffix == '.csv':
        return _read_rows_csv(path)
    raise ValueError(f'Неподдерживаемый формат файла: {suffix or "без расширения"}')


def preview_headers(file_path: str, header_row: int) -> List[str]:
    rows = read_sheet_rows(file_path)
    for row_number, cells in rows:
        if row_number == header_row:
            return [cell_to_str(c) for c in cells]
    return []


def iter_price_rows(
    file_path: str,
    header_row: int,
    data_start_row: int,
    column_map: dict,
) -> Iterator[ParsedPriceRow]:
    rows = read_sheet_rows(file_path)
    header_cells: List[str] = []
    for row_number, cells in rows:
        if row_number == header_row:
            header_cells = [cell_to_str(c) for c in cells]
            break
    indexes = build_column_indexes(column_map, header_cells)

    article_idx = indexes.get('article')
    price_idx = indexes.get('price')
    if article_idx is None or price_idx is None:
        raise ValueError('Укажите колонки для артикула и цены')

    brand_idx = indexes.get('brand')
    qty_idx = indexes.get('qty')
    external_idx = indexes.get('external_id')

    for row_number, cells in rows:
        if row_number < data_start_row:
            continue

        def get_cell(idx: int) -> str:
            if idx is None or idx >= len(cells):
                return ''
            return cell_to_str(cells[idx])

        article = get_cell(article_idx)
        if not article:
            continue

        price = parse_price(get_cell(price_idx))
        if price is None:
            continue

        yield ParsedPriceRow(
            row_number=row_number,
            article=article,
            brand=get_cell(brand_idx) if brand_idx is not None else '',
            price=price,
            quantity=parse_quantity(get_cell(qty_idx)) if qty_idx is not None else 0,
            external_id=get_cell(external_idx) if external_idx is not None else '',
        )
